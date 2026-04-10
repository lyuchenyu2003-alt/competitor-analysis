"""
竞品内容分析工具
新增：动态粉丝分层 · 品类嵌套词库 · 高分辨率特征 · Excel 图表导出
"""

import sys, importlib, io, os, re, json
from collections import Counter
from functools import lru_cache
from datetime import datetime

# ══════════════════════════════════════════════════════════════
# 依赖检查
# ══════════════════════════════════════════════════════════════
_MISSING = [p for p in ["jieba", "openpyxl", "plotly"]
            if importlib.util.find_spec(p) is None]

import streamlit as st
st.set_page_config(page_title="竞品内容分析工具 · Pro", page_icon="📊", layout="wide")

if _MISSING:
    st.error("缺少必要依赖，请在终端运行：")
    for p in _MISSING:
        st.code(f"pip3 install {p}")
    st.stop()

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import jieba

# ══════════════════════════════════════════════════════════════
# 文本分析与关键词提取 (恢复辅助函数)
# ══════════════════════════════════════════════════════════════
STOP_WORDS = frozenset({
    "的", "了", "和", "是", "在", "我", "也", "这", "有", "不",
    "都", "就", "但", "让", "用", "来", "能", "还", "被", "把",
    "啊", "吧", "呢", "哦", "哈", "嗯", "个", "一", "这个", "那个",
})

@lru_cache(maxsize=4096)
def _tokenize(title: str) -> tuple:
    return tuple(w for w in jieba.cut(str(title))
                 if len(w) >= 2 and w not in STOP_WORDS)

def extract_keywords(titles: list, top_n: int = 20) -> list:
    words = [w for t in titles for w in _tokenize(str(t))]
    return Counter(words).most_common(top_n)

try:
    from openai import OpenAI as _OpenAI
    _HAS_OPENAI = True
except ImportError:
    _HAS_OPENAI = False

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    _HAS_OPX = True
except ImportError:
    _HAS_OPX = False

import sys, importlib, io, os, re, json
from collections import Counter
from functools import lru_cache
from datetime import datetime
import pandas as pd
import numpy as np

# ══════════════════════════════════════════════════════════════
# 常量 & 模糊匹配別名字典
# ══════════════════════════════════════════════════════════════
COLUMN_ALIASES = {
    "品牌名称":     ["品牌", "brand", "品牌名"],
    "内容标题":     ["标题", "title", "内容", "文章标题"],
    "内容正文":     ["正文", "content", "文章正文", "文案"], # 新增：用于内容关键字识别
    "发布平台":     ["平台", "platform", "渠道"],
    "发布日期":     ["日期", "date"],
    "发布时间":     ["时间", "time"],
    "点赞数":       ["点赞", "likes", "赞数"],
    "收藏数":       ["收藏", "favorites", "收藏量"],
    "评论数":       ["评论", "comments"],
    "分享数":       ["分享", "shares", "转发数"],
    "博主粉丝量":   ["粉丝量", "粉丝数", "followers", "粉丝"],
    "博主类型":     ["博主级别", "kol类型", "账号类型"],
    "品类":         ["类目", "category", "产品类别"],
    "营销动机":     ["类别", "动机", "marketing_motivation"], # 字段更名适配
    # 多维度支持新增
    "粉丝年龄段":   ["年龄段", "audience_age", "粉丝年龄"],
    "核心受众画像": ["受众画像", "audience", "用户画像", "群体类型"],
    "二级内容类型": ["子内容类型", "content_subtype"],
    "二级博主身份": ["博主子类型", "blogger_subtype", "细分身份"],
}

REQUIRED_COLS = ["品牌名称", "内容标题", "点赞数", "收藏数", "评论数", "分享数"]
# 注意：移除了硬编码的层级标签，由动态计算生成

# ══════════════════════════════════════════════════════════════
# 模块一：双轨语料库管理 (通用标题钩子 vs 行业特定词汇)
# ══════════════════════════════════════════════════════════════
DEFAULT_GENERAL_HOOKS = {
    "痛点焦虑": ["避雷", "千万别", "后悔", "救命", "踩坑", "警惕", "别买", "翻车", "智商税"],
    "利益获得": ["保姆级", "平替", "干货", "省下", "清单", "攻略", "合集", "教程", "免费", "一单解决"],
    "权威背书": ["亲测", "配方", "成分党", "实验室", "研究", "临床", "大牌同款"],
    "社交货币": ["天花板", "氛围感", "私藏", "封神", "绝了", "顶级", "神器", "爆款", "绝绝子"]
}

DEFAULT_INDUSTRY_KWS = {
    "护肤": {
        "痛点焦虑": ["红血丝", "爆痘", "闷痘", "烂脸", "过敏", "脱皮", "敏感肌"],
        "权威背书": ["皮肤科医生", "早C晚A", "烟酰胺", "玻色因"],
    },
    "食品饮料": {
        "痛点焦虑": ["减脂", "低卡", "无糖", "戒糖", "戒油", "热量", "升糖"],
        "权威背书": ["营养师", "配料表", "0添加", "低GI"],
    }
}

def df_to_lexicon(df: pd.DataFrame, is_industry: bool = False) -> dict:
    """DataFrame 转 嵌套字典格式"""
    lexicon = {}
    for _, row in df.iterrows():
        motive = str(row.get("营销动机", "")).strip()
        kw = str(row.get("关键词", "")).strip()
        if not motive or not kw or kw == "nan": continue
        
        if is_industry:
            cat = str(row.get("品类", "")).strip()
            if not cat: continue
            lexicon.setdefault(cat, {}).setdefault(motive, [])
            if kw not in lexicon[cat][motive]: lexicon[cat][motive].append(kw)
        else:
            lexicon.setdefault(motive, [])
            if kw not in lexicon[motive]: lexicon[motive].append(kw)
    return lexicon

# ══════════════════════════════════════════════════════════════
# 模块二：高分辨率特征与动机识别
# ══════════════════════════════════════════════════════════════
def classify_title_length_hires(title: str) -> str:
    """高分辨率特征：标题长度 10字区间分类"""
    n = len(str(title).strip())
    if n == 0: return "[0]"
    low = ((n - 1) // 10) * 10 + 1
    return f"[{low}-{low + 9}]"

def detect_motivation_dual_track(title: str, content: str, category: str, general_lex: dict, industry_lex: dict) -> str:
    """
    双轨逻辑：
    1. 标题关键匹配（通用模版库）优先
    2. 内容/标题匹配（行业特有库）辅助补充
    """
    t = str(title) if pd.notna(title) else ""
    c = str(content) if pd.notna(content) else ""
    text_pool = t + " " + c
    
    matched_motives = set()
    
    # 轨道 1：通用标题钩子
    for motive, kws in general_lex.items():
        if any(kw in t for kw in kws):
            matched_motives.add(motive)
            
    # 轨道 2：行业内容词汇
    cat_lex = industry_lex.get(category, {})
    for motive, kws in cat_lex.items():
        if any(kw in text_pool for kw in kws):
            matched_motives.add(motive)
            
    if not matched_motives: return "无明显动机"
    if len(matched_motives) == 1: return list(matched_motives)[0]
    return "复合策略"

def compute_dynamic_fan_tiers(fans: pd.Series) -> tuple:
    """动态粉丝分层 (pd.qcut 四分位数)"""
    fans_clean = fans.clip(lower=0)
    unique_nonzero = fans_clean[fans_clean > 0].nunique()

    labels = ["头部(Top 25%)", "头腰(25%-50%)", "腰部(50%-75%)", "草根(Bottom 25%)"]
    
    if unique_nonzero < 4:
        # 数据极度缺乏时的降级方案
        return pd.Series(["无法计算分位"] * len(fans_clean)), None
        
    try:
        # qcut 自动划分为4个均等数量的区间
        tiers, bins = pd.qcut(
            fans_clean.clip(lower=1), 
            q=4, 
            labels=labels[::-1], # 倒序，数值越小越是草根
            retbins=True, 
            duplicates="drop"
        )
        return tiers.astype(str), [float(b) for b in bins]
    except Exception:
        return pd.Series(["异常分层"] * len(fans_clean)), None

# 数值清洗 & 格式化 (恢复辅助函数)
# ══════════════════════════════════════════════════════════════
def clean_number(val) -> float:
    """支持 '1.5w', '1.5W', '1.5万', '3.2k', '1,000' 等非标格式"""
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace("\uff0c", ",")
    m = re.match(r"^([\d.]+)[\u4e07wW]$", s)      # 万/w/W
    if m:
        return float(m.group(1)) * 10000
    m = re.match(r"^([\d.]+)[kK]$", s)             # k/K
    if m:
        return float(m.group(1)) * 1000
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0

def fmt_num(n) -> str:
    """大数字格式化：>= 1w 显示 Xw，>= 1k 显示 Xk"""
    if pd.isna(n):
        return "-"
    n = float(n)
    if n >= 10000:
        return f"{n / 10000:.1f}w"
    if n >= 1000:
        return f"{n / 1000:.1f}k"
    return f"{n:.0f}"

# 👇👇👇 贴入这段关键词提取辅助函数 👇👇👇
# ══════════════════════════════════════════════════════════════
# 文本分析与关键词提取 (恢复辅助函数)
# ══════════════════════════════════════════════════════════════
STOP_WORDS = frozenset({
    "的", "了", "和", "是", "在", "我", "也", "这", "有", "不",
    "都", "就", "但", "让", "用", "来", "能", "还", "被", "把",
    "啊", "吧", "呢", "哦", "哈", "嗯", "个", "一", "这个", "那个",
})

@lru_cache(maxsize=4096)
def _tokenize(title: str) -> tuple:
    return tuple(w for w in jieba.cut(str(title))
                 if len(w) >= 2 and w not in STOP_WORDS)

def extract_keywords(titles: list, top_n: int = 20) -> list:
    words = [w for t in titles for w in _tokenize(str(t))]
    return Counter(words).most_common(top_n)
# 👆👆👆 贴入结束 👆👆👆

# ══════════════════════════════════════════════════════════════
# 模块三：核心数据加载与预处理 (包含报错防范与黑马权重)
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner="⚙️ 深度解析处理中，请稍候…")
def load_data(file_bytes: bytes, file_name: str, 
              general_lex_json: str, industry_lex_json: str,
              dh_weight_eff: float = 0.7, dh_weight_scale: float = 0.3,
              enable_dewater: bool = False, rates_json: str = "{}") -> tuple:
    """
    返回: (df处理结果 | None, 缺失列预检报错列表, 分层bins | None)
    新增参数: dh_weight_eff (效率权重), dh_weight_scale (规模权重)
    """
    general_lex = json.loads(general_lex_json)
    industry_lex = json.loads(industry_lex_json)

    if not file_name.lower().endswith(".xlsx"):
        return None, ["文件格式错误：请使用 .xlsx 格式。"], None

    df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    
    # 模糊匹配列名
    rename_map = {}
    existing_cols = {c.lower(): c for c in df.columns}
    for standard, aliases in COLUMN_ALIASES.items():
        if standard in df.columns: continue
        for alias in aliases:
            if alias.lower() in existing_cols:
                rename_map[existing_cols[alias.lower()]] = standard
                break
    df = df.rename(columns=rename_map)

    # 预检机制：如果缺失必需列，不报错，直接返回清单给前端显示
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        return None, [f"缺失必需列：{'、'.join(missing)}。请检查模板。"], None

    # 数值列安全清洗
    num_cols = ["点赞数", "收藏数", "评论数", "分享数", "博主粉丝量"]
    for col in num_cols:
        if col in df.columns:
            # 引入你之前的 clean_number 逻辑 (1.2w -> 12000)
            df[col] = df[col].apply(clean_number) 
        else:
            df[col] = 0.0

    # 核心业务逻辑：双轨语料识别 & 人工优先
    has_cat = "品类" in df.columns
    has_content = "内容正文" in df.columns
    
    # 【人工标注优先逻辑】：如果Excel中已有「营销动机」列，且有值，则不被规则覆盖
    if "营销动机" not in df.columns:
        df["营销动机"] = None
        
    df["机器识别_营销动机"] = df.apply(
        lambda r: detect_motivation_dual_track(
            title=r["内容标题"],
            content=r["内容正文"] if has_content else "",
            category=str(r["品类"]) if has_cat else "通用",
            general_lex=general_lex,
            industry_lex=industry_lex
        ), axis=1
    )
    # 人工标注(原值)为空时，填入机器识别结果
    df["营销动机"] = df["营销动机"].replace("", np.nan).fillna(df["机器识别_营销动机"])
    # 高分辨率特征：标题长度
    df["标题长度区间"] = df["内容标题"].apply(classify_title_length_hires)

    # 👇👇👇 新增：执行跨平台去水折算 👇👇👇
    if enable_dewater:
        rates = json.loads(rates_json)
        plat_col = "发布平台" if "发布平台" in df.columns else None
        
        if plat_col:
            # 匹配对应系数，找不到的默认 1.0 (不折算)
            df["_水位系数"] = df[plat_col].map(rates).fillna(1.0)
            
            # 对所有【绝对规模指标】进行等比例折算
            # 注意：粉丝量和互动量一起折算，这样能保证基础的互动率 (%) 保持真实不变，
            # 但能让小红书和抖音在散点图的 X 轴（发布量/点赞规模）上做到公平对齐！
            for col in ["点赞数", "收藏数", "评论数", "分享数", "博主粉丝量"]:
                if col in df.columns:
                    df[col] = df[col] * df["_水位系数"]
    # 👆👆👆 新增结束 👆👆👆

    # 互动率计算
    df["互动总量"] = df["点赞数"] + df["收藏数"] + df["评论数"] + df["分享数"]
    # 互动率计算
    df["互动总量"] = df["点赞数"] + df["收藏数"] + df["评论数"] + df["分享数"]
    df["互动率"] = np.where(df["博主粉丝量"] > 0, round(df["互动总量"] / df["博主粉丝量"], 6), 0.0)

# 动态粉丝分层 (核心变更) - 恢复 UI 期望的列名
    df["粉丝层级"], fan_bins = compute_dynamic_fan_tiers(df["博主粉丝量"])

    # 动态黑马指数计算 - 把中间计算结果存回 DataFrame 供 UI 渲染图表
    df["层级平均互动率"] = df.groupby("粉丝层级")["互动率"].transform('mean').fillna(0)
    df["品牌平均点赞"] = df.groupby("品牌名称")["点赞数"].transform('mean').fillna(1)
    
    # 效率(互动率偏离度)与规模(点赞偏离度)
    efficiency_score = np.where(df["层级平均互动率"] > 0, df["互动率"] / df["层级平均互动率"], 0)
    scale_score = np.where(df["品牌平均点赞"] > 0, df["点赞数"] / df["品牌平均点赞"], 0)
    
# 根据用户调节的滑块权重进行计算
    df["黑马指数"] = np.round(efficiency_score * dh_weight_eff + scale_score * dh_weight_scale, 3)

    # ================= 时间特征处理 =================
    if "发布日期" in df.columns:
        df["发布日期"] = pd.to_datetime(df["发布日期"], errors="coerce")
    else:
        df["发布日期"] = pd.NaT

    if "发布时间" in df.columns:
        df["发布小时"] = pd.to_numeric(df["发布时间"].astype(str).str[:2], errors="coerce").fillna(12).astype(int)
    else:
        df["发布小时"] = 12

    df["发布星期"] = df["发布日期"].dt.day_name()
    # ===============================================

    return df, [], fan_bins
# ══════════════════════════════════════════════════════════════
# Strategy Facts
# ══════════════════════════════════════════════════════════════
def build_strategy_facts(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for brand, g in df.groupby("品牌名称"):
        hour_g  = g.groupby("发布小时")["互动总量"].mean()
        best_h  = int(hour_g.idxmax()) if not hour_g.empty else -1
        mot_g   = g.groupby("营销动机")["互动总量"].mean()
        top_mot = mot_g.idxmax() if not mot_g.empty else "无"
        top_row = g.loc[g["黑马指数"].idxmax()]
        b_mode  = (g["博主类型"].mode() if "博主类型" in g.columns
                   else pd.Series(["未知"]))
        records.append({
            "品牌":              brand,
            "内容总量":           len(g),
            "平均互动率":          round(g["互动率"].mean(), 6),
            "平均黑马指数":         round(g["黑马指数"].mean(), 3),
            "黑马内容占比":         f"{(g['黑马指数'] > 1.2).mean():.1%}",
            "最高黑马指数":         round(g["黑马指数"].max(), 3),
            "最高黑马内容标题":      top_row["内容标题"],
            "最佳发布时段(h)":      best_h,
            "高互动营销动机":        top_mot,
            "主力博主类型":          b_mode.iloc[0] if not b_mode.empty else "未知",
            "平均点赞":             fmt_num(g["点赞数"].mean()),
            "平均收藏":             fmt_num(g["收藏数"].mean()),
        })
    return pd.DataFrame(records)

# ══════════════════════════════════════════════════════════════
# Excel 导出（含图表 / 首行冻结 / 自动筛选 / 色阶 / Methodology）
# ══════════════════════════════════════════════════════════════
def build_excel(flt, strategy_facts, top10, kw_df, brand_df, ai_insight: str,
                interact_df=None, brand_top5_df=None, platform_df=None) -> io.BytesIO:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 导出时去掉机器生成的中间列，保持表格整洁
        export_drop = [c for c in ["机器识别_营销动机"] if c in flt.columns]
        flt_export = flt.drop(columns=export_drop)
        flt_export.to_excel(writer,     sheet_name="原始数据",       index=False)
        top10.to_excel(writer,          sheet_name="TOP10爆款内容",  index=False)
        kw_df.to_excel(writer,          sheet_name="关键词分析",     index=False)
        brand_df.to_excel(writer,       sheet_name="竞品对比",       index=False)
        strategy_facts.to_excel(writer, sheet_name="Strategy_Facts", index=False)

        # 新增：互动类型拆解 / 品牌 TOP5 / 平台对比 sheets
        if interact_df is not None and len(interact_df) > 0:
            interact_df.to_excel(writer, sheet_name="互动类型拆解", index=False)
        if brand_top5_df is not None and len(brand_top5_df) > 0:
            brand_top5_df.to_excel(writer, sheet_name="品牌TOP5", index=False)
        if platform_df is not None and len(platform_df) > 0:
            platform_df.to_excel(writer, sheet_name="平台效果对比", index=False)

        # 1. 自动插入 AI 洞察页（美化版：按 ## / ### 拆分段落写入多行）
        insight_text = ai_insight or "(尚未生成 AI 洞察报告)"
        insight_lines = []
        for line in insight_text.split("\n"):
            stripped = line.strip()
            if stripped:
                insight_lines.append(stripped)
        if not insight_lines:
            insight_lines = [insight_text]
        insight_rows = [{"行号": i + 1, "AI 洞察报告内容": l} for i, l in enumerate(insight_lines)]
        pd.DataFrame(insight_rows).to_excel(writer, sheet_name="AI_Insight", index=False)

        # 2. 自动插入 Methodology 指标逻辑说明页
        pd.DataFrame({
            "指标名称": ["互动率", "动态粉丝分层", "黑马指数"],
            "业务逻辑说明": [
                "互动总量 / 博主粉丝量。衡量内容在粉丝群体中的穿透效率。",
                "根据本次上传数据的实际分布，使用四分位数(qcut)动态划分为草根、腰部、头腰、头部。",
                "黑马内容判定标准：综合考虑『互动率偏离度』与『点赞规模偏离度』，大于1.2视为黑马。",
            ]
        }).to_excel(writer, sheet_name="Methodology", index=False)

        if not _HAS_OPX:
            return output

        wb = writer.book
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.chart import ScatterChart, Reference, Series
        
        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        hdr_font   = Font(bold=True, color="FFFFFF", size=11)
        center     = Alignment(horizontal="center", vertical="center")
        thin       = Side(style="thin")
        bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 3. 全局样式：首行冻结 + 自动筛选 + 深蓝表头 + 列宽自适应
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, bdr
            for col_cells in ws.columns:
                ws.column_dimensions[col_cells[0].column_letter].width = 16

        # 4. Strategy_Facts：黑马指数列开启三色色阶 (红-黄-绿)
        sf_ws = wb["Strategy_Facts"]
        bhi_col = None
        for idx, cell in enumerate(sf_ws[1], start=1):
            if cell.value == "平均黑马指数":
                bhi_col = idx; break
                
        if bhi_col:
            col_l = get_column_letter(bhi_col)
            sf_ws.conditional_formatting.add(
                f"{col_l}2:{col_l}{sf_ws.max_row}",
                ColorScaleRule(start_type="min", start_color="FFC7CE", 
                               mid_type="percentile", mid_value=50, mid_color="FFEB9C", 
                               end_type="max", end_color="C6EFCE")
            )

        # 5. 竞品对比 Sheet：插入四象限策略散点图 (修复连线与品牌图例)
        cmp_ws = wb["竞品对比"]
        if cmp_ws.max_row > 1:
            chart = ScatterChart()
            chart.title = "品牌策略四象限 (X=发布量, Y=互动率)"
            chart.x_axis.title = "内容发布量"
            chart.y_axis.title = "平均互动率"
            chart.width, chart.height = 18, 12
            chart.style = 13 # 设定为优雅的默认散点图样式
            
            # 遍历每一行，将每个品牌作为【独立的数据系列】加入图表
            for i in range(2, cmp_ws.max_row + 1):
                brand_name = cmp_ws.cell(row=i, column=1).value
                # X 轴引用第 2 列(发布量)，Y 轴引用第 7 列(互动率)
                x_val = Reference(cmp_ws, min_col=2, min_row=i, max_row=i)
                y_val = Reference(cmp_ws, min_col=7, min_row=i, max_row=i)
                
                series = Series(values=y_val, xvalues=x_val, title=brand_name)
                
                # 关键修复：强制设定为圆点，并且【关闭连线】
                series.marker.symbol = "circle"
                series.graphicalProperties.line.noFill = True 
                
                chart.series.append(series)
                
            cmp_ws.add_chart(chart, "I2")

        # 6. AI_Insight Sheet 美化：标题行高亮 + 自动换行 + 加宽列
        if "AI_Insight" in wb.sheetnames:
            ai_ws = wb["AI_Insight"]
            ai_ws.column_dimensions["A"].width = 8
            ai_ws.column_dimensions["B"].width = 80
            wrap_align = Alignment(wrap_text=True, vertical="top")
            section_fill = PatternFill("solid", fgColor="E8F0FE")
            section_font = Font(bold=True, size=12, color="1F4E79")
            normal_font  = Font(size=11, color="333333")
            for row in ai_ws.iter_rows(min_row=2, max_row=ai_ws.max_row, min_col=2, max_col=2):
                cell = row[0]
                cell.alignment = wrap_align
                text_val = str(cell.value or "")
                if text_val.startswith("#") or text_val.startswith("**"):
                    cell.fill = section_fill
                    cell.font = section_font
                    ai_ws.row_dimensions[cell.row].height = 28
                else:
                    cell.font = normal_font
                    ai_ws.row_dimensions[cell.row].height = 22

    return output
# ══════════════════════════════════════════════════════════════
# AI 报告生成
# ══════════════════════════════════════════════════════════════
def format_prompt(facts_df: pd.DataFrame, flt_df: pd.DataFrame = None) -> str:
    text = facts_df.to_string(index=False)
    if len(text) > 2500:
        text = text[:2500] + "\n...[截断]"

    # Build supplementary data summaries if flt_df is available
    extra_sections = ""
    if flt_df is not None and len(flt_df) > 0:
        # Interaction type breakdown
        interact_cols = ["点赞数", "收藏数", "评论数", "分享数"]
        avail = [c for c in interact_cols if c in flt_df.columns]
        if avail:
            sums = {c: int(flt_df[c].sum()) for c in avail}
            extra_sections += "\n【互动类型拆解】\n" + "\n".join(f"- {k}: {v:,}" for k, v in sums.items())

        # Brand TOP 5
        brand5 = flt_df.groupby("品牌名称")["互动总量"].sum().nlargest(5)
        extra_sections += "\n\n【品牌互动 TOP 5】\n" + "\n".join(
            f"- {b}: {int(v):,}" for b, v in brand5.items()
        )

        # Platform comparison
        if "发布平台" in flt_df.columns:
            plat = flt_df.groupby("发布平台").agg(
                avg_interact=("互动总量", "mean"),
                count=("内容标题", "count")
            ).round(1)
            extra_sections += "\n\n【发布平台效果】\n" + "\n".join(
                f"- {p}: 平均互动={row['avg_interact']:.0f}, 发布量={int(row['count'])}"
                for p, row in plat.iterrows()
            )

        # Posting time pattern
        if "发布小时" in flt_df.columns:
            top_hours = flt_df.groupby("发布小时")["互动总量"].mean().nlargest(3)
            extra_sections += "\n\n【最佳发布时段 TOP 3】\n" + "\n".join(
                f"- {int(h)}:00 平均互动={v:.0f}" for h, v in top_hours.items()
            )

    return (
        "你是一位资深内容营销分析师，擅长从数据中提炼可落地的策略洞察。\n"
        "请基于以下竞品数据，撰写一份结构清晰的分析报告。\n\n"
        "## 报告结构（请严格按以下六个板块输出）\n\n"
        "### 一、数据概览\n"
        "简述本次分析的数据规模（品牌数、内容数、时间跨度）及整体互动水平。\n\n"
        "### 二、品牌竞争力排名\n"
        "结合互动总量与黑马指数，点评 TOP 3 品牌的策略特征；指出最具增长潜力的品牌。\n\n"
        "### 三、互动结构洞察\n"
        "分析点赞/收藏/评论/分享的比例关系，判断该品类用户的互动偏好特征。\n\n"
        "### 四、内容策略洞察\n"
        "哪种营销动机(成分科普/真实测评/场景种草等)效果最好？黑马内容有什么共性特征？最佳标题长度区间是多少字？\n\n"
        "### 五、渠道与时段洞察\n"
        "各发布平台效果差异；最佳发布时段和星期建议。\n\n"
        "### 六、可执行建议（3-5 条）\n"
        "给出具体、可直接落地的内容策略建议，每条注明适用的品牌类型或场景。\n\n"
        "---以下为数据---\n"
        "【Strategy Facts 竞品汇总表】\n"
        f"{text}\n"
        f"{extra_sections}\n"
        "---数据结束---\n\n"
        "要求：中文输出，800 字以内，语言专业但易读，"
        "多用数据佐证结论，避免空洞描述。每个板块 2-4 句话即可。"
    )


# 国内可用大模型配置表（均兼容 OpenAI SDK 格式）
MODEL_CONFIGS = {
    "Kimi":    {"base_url": "https://api.moonshot.cn/v1",
                "model":    "moonshot-v1-8k"},
    "DeepSeek":{"base_url": "https://api.deepseek.com",
                "model":    "deepseek-chat"},
    "通义千问": {"base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
                "model":    "qwen-turbo"},
    "豆包":    {"base_url": "https://ark.cn-beijing.volces.com/api/v3",
                "model":    "doubao-pro-32k"},
}

def call_ai(api_key: str, model_type: str, prompt: str, temperature: float = 0.7) -> str:
    if not _HAS_OPENAI:
        return "请先安装 openai 库：pip3 install openai"
    cfg    = MODEL_CONFIGS.get(model_type, MODEL_CONFIGS["Kimi"])
    client = _OpenAI(api_key=api_key, base_url=cfg["base_url"])
    resp   = client.chat.completions.create(
        model=cfg["model"],
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1200,
        temperature=temperature,
        timeout=60
    )
    return resp.choices[0].message.content

# ══════════════════════════════════════════════════════════════
# 品类词库本地化管理 (恢复辅助函数)
# ══════════════════════════════════════════════════════════════
CATEGORY_LEXICON_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "category_lexicon.json")

# ── 本地配置：自动读取 Kimi 默认 Key ────────────────────────────
_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
def _load_config() -> dict:
    try:
        with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}
_CONFIG = _load_config()
KIMI_DEFAULT_KEY = _CONFIG.get("kimi_api_key", "")

def load_category_lexicon() -> dict:
    if os.path.exists(CATEGORY_LEXICON_PATH):
        try:
            with open(CATEGORY_LEXICON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # 兼容新的双轨词库：合并通用与行业默认词库
    merged = {"通用": DEFAULT_GENERAL_HOOKS}
    merged.update(DEFAULT_INDUSTRY_KWS)
    return merged

def save_category_lexicon(lexicon: dict) -> None:
    with open(CATEGORY_LEXICON_PATH, "w", encoding="utf-8") as f:
        json.dump(lexicon, f, ensure_ascii=False, indent=2)

def nested_lexicon_to_df(lexicon: dict) -> pd.DataFrame:
    rows = [
        {"品类": cat, "动机": motive, "关键词": kw}
        for cat, motives in lexicon.items()
        for motive, kws in motives.items()
        for kw in kws
    ]
    return pd.DataFrame(rows, columns=["品类", "动机", "关键词"])

def df_to_nested_lexicon(df: pd.DataFrame) -> dict:
    lexicon = {}
    for _, row in df.iterrows():
        cat    = str(row.get("品类", "")).strip()
        motive = str(row.get("动机", "")).strip()
        kw     = str(row.get("关键词", "")).strip()
        if cat and motive and kw and kw != "nan":
            lexicon.setdefault(cat, {}).setdefault(motive, [])
            if kw not in lexicon[cat][motive]:
                lexicon[cat][motive].append(kw)
    return lexicon


# ══════════════════════════════════════════════════════════════
# Session State 初始化
# ══════════════════════════════════════════════════════════════
if "category_lexicon" not in st.session_state:
    st.session_state.category_lexicon = load_category_lexicon()
if "ai_insight" not in st.session_state:
    st.session_state.ai_insight = ""
if "fan_bins" not in st.session_state:
    st.session_state.fan_bins = None


# ══════════════════════════════════════════════════════════════
# 侧边栏
# ══════════════════════════════════════════════════════════════
# --- 数据上传与示例加载区 ---
st.sidebar.header("📂 数据源配置")

# 1. 增加一个勾选框：使用示例数据
use_sample = st.sidebar.checkbox("使用美妆护肤示例数据", value=False, help="开启后将自动加载 500 条行业仿真数据，无需手动上传")

uploaded_file = st.sidebar.file_uploader("或者上传你自己的 Excel 文件", type=["xlsx"])

# 2. 核心逻辑切换
df = None
file_bytes = None
file_name = ""

if use_sample:
    # 自动读取本地的示例文件
    sample_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Skincare_Marketing_Final_Demo.xlsx")
    try:
        with open(sample_path, "rb") as f:
            file_bytes = f.read()
        file_name = sample_path
        st.sidebar.success("✅ 已加载行业示例数据")
    except FileNotFoundError:
        st.sidebar.error("❌ 未找到示例文件，请确保文件在代码目录下")
elif uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()
    file_name = uploaded_file.name
    st.sidebar.success(f"✅ 已上传: {file_name}")

# ── 侧边栏：AI 配置 + BHI 权重 + 去水 + 词库（始终渲染）──────
with st.sidebar:
    st.markdown("---")
    st.header("🤖 AI 配置")
    model_choice = st.selectbox("模型", [
        "Kimi（月之暗面·推荐）",
        "DeepSeek（深度求索·便宜）",
        "通义千问（阿里云）",
        "豆包（字节跳动）",
    ])
    model_type   = model_choice.split("（")[0]
    _default_key = KIMI_DEFAULT_KEY if model_type == "Kimi" else ""
    _label       = "API Key（已预设）" if model_type == "Kimi" else "API Key"
    _placeholder = "Kimi Key 已自动载入" if model_type == "Kimi" else "请输入对应模型的 API Key"
    api_key = st.text_input(
        _label, value=_default_key, type="password",
        placeholder=_placeholder, key=f"apikey_{model_type}",
    )
    if not _HAS_OPENAI:
        st.warning("未检测到 openai 库，请在终端运行：\npip3 install openai")
    ai_temperature = st.slider(
        "🌡️ 创意温度",
        min_value=0.8, max_value=1.0, value=0.9, step=0.05,
        help="0.80 = 贴近数据、稳健严谨；1.00 = 发散创意、风格多变（Kimi 仅支持 0.8-1.0）",
        key="ai_temperature",
    )

    st.markdown("---")
    st.header("🎛️ 黑马指数调节")
    dh_weight_eff   = st.slider("效率权重（互动率）", 0.0, 1.0, 0.7, 0.1, key="weight_slider")
    dh_weight_scale = 1.0 - dh_weight_eff
    st.caption(f"当前规模权重（点赞数）为: {dh_weight_scale:.1f}")

    st.markdown("---")
    st.header("🌊 跨平台水位折算")
    enable_dewater = st.toggle("开启全域去水模式", value=False,
                               help="开启后，所有流量规模指标将乘以对应系数，折算为统一基准（消除通胀）")
    default_rates = pd.DataFrame([
        {"平台": "小红书",    "水份系数": 1.0},
        {"平台": "抖音",      "水份系数": 0.1},
        {"平台": "微信视频号", "水份系数": 2.0},
        {"平台": "B站",       "水份系数": 1.5},
        {"平台": "微博",      "水份系数": 0.05},
    ])
    edited_rates = st.data_editor(
        default_rates, num_rows="dynamic", hide_index=True,
        use_container_width=True, disabled=not enable_dewater,
    )
    rates_dict = dict(zip(edited_rates["平台"], edited_rates["水份系数"]))
    rates_json = json.dumps(rates_dict, ensure_ascii=False)



# ══════════════════════════════════════════════════════════════
# 主界面
# ══════════════════════════════════════════════════════════════
st.title("📊 竞品内容分析工具")
st.caption("动态粉丝分层 · 黑马指数 · 跨平台去水 · AI 洞察 · Excel 图表导出")

if not file_bytes:
    st.info("👈 请在左侧上传 Excel 数据文件，或勾选「使用美妆护肤示例数据」开始分析")

    with st.expander("🐴 黑马指数 (BHI) 说明", expanded=True):
        st.markdown("""
**什么是黑马指数？** 用来发现"以小博大"的内容 — 粉丝量不高但互动远超同层级平均水平。

**公式**：BHI = 互动率偏离度 x 效率权重 + 点赞规模偏离度 x 规模权重

BHI > 1.2 即为黑马内容，说明该内容的互动表现显著超过同粉丝层级的平均水平。左侧边栏可调节效率/规模的权重比例。
        """)

    with st.expander("🌊 跨平台去水折算说明"):
        st.markdown("""
**为什么要去水？** 不同平台的流量"含水量"不同。例如抖音一条视频动辄10w赞，但小红书1w赞已是大爆款。直接对比原始数据会产生严重误判。

开启去水模式后，所有互动数据会乘以对应平台的折算系数（以小红书为基准 1.0），消除平台间的通胀差异，实现真正的跨平台可比。
        """)

    with st.expander("📊 动态粉丝分层说明"):
        st.markdown("""
基于当前上传数据的实际粉丝分布，使用四分位数 (pd.qcut) 动态划分为四层：草根、腰部、头腰、头部。每层占 25%，阈值随数据自动变化，上传后在页面顶部展示具体分层边界。
        """)

else:
    # file_bytes / file_name 已在侧边栏数据源区设置好（支持示例数据和上传文件）
    lexicon_dict = st.session_state.category_lexicon
    general_lex  = lexicon_dict.get("通用", {})
    industry_lex = {k: v for k, v in lexicon_dict.items() if k != "通用"}
    gen_json = json.dumps(general_lex,  ensure_ascii=False)
    ind_json = json.dumps(industry_lex, ensure_ascii=False)

    df, errors, fan_bins = load_data(
        file_bytes, file_name,
        gen_json, ind_json,
        dh_weight_eff, dh_weight_scale,
        enable_dewater, rates_json,
    )

    if errors:
        # 不崩溃：显示模板补全建议
        st.error("📋 数据检查未通过，请参考以下建议修正你的 Excel：")
        for err in errors:
            st.error(err)

        st.warning("**模板补全建议**：下载标准模板，对照填写后重新上传")
        template_buf = io.BytesIO()
        pd.DataFrame(columns=REQUIRED_COLS + OPTIONAL_COLS).to_excel(
            template_buf, index=False, engine="openpyxl"
        )
        st.download_button(
            "📥 下载标准数据模板",
            data=template_buf.getvalue(),
            file_name="竞品分析数据模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.stop()

    st.session_state.fan_bins = fan_bins
    st.success(
        f"✅ 数据加载成功！共 **{len(df)}** 条记录，"
        f"**{df['品牌名称'].nunique()}** 个品牌，特征工程已完成"
    )

    # 动态分层阈值展示
    if fan_bins and len(fan_bins) >= 4:
        with st.expander("📊 当前数据粉丝分层阈值（动态 · 基于 pd.qcut）", expanded=True):
            cols = st.columns(4)
            tier_desc = [
                ("草根", f"< {fmt_num(fan_bins[1])} 粉"),
                ("腰部", f"{fmt_num(fan_bins[1])} - {fmt_num(fan_bins[2])} 粉"),
                ("头腰", f"{fmt_num(fan_bins[2])} - {fmt_num(fan_bins[3])} 粉"),
                ("头部", f"> {fmt_num(fan_bins[3])} 粉"),
            ]
            for col_ui, (tier, rng) in zip(cols, tier_desc):
                col_ui.metric(tier, rng,
                              help=f"基于当前上传数据的四分位分层，非固定阈值")
    else:
        st.info("粉丝数据唯一值不足，已退化为固定阈值（1w / 5w / 20w）")

    # 筛选器（含二级可选字段）
    with st.expander("🔍 数据筛选", expanded=True):
        c1, c2, c3 = st.columns(3)
        sel_brand = c1.selectbox("品牌", ["全部"] + sorted(df["品牌名称"].unique().tolist()))
        sel_cat   = c2.selectbox(
            "品类", ["全部"] + (sorted(df["品类"].unique().tolist()) if "品类" in df.columns else [])
        )
        sel_plat  = c3.selectbox(
            "平台", ["全部"] + (sorted(df["发布平台"].unique().tolist()) if "发布平台" in df.columns else [])
        )

        # 二级筛选（按实际字段存在动态显示）
        extra_available = [c for c in ["粉丝年龄段", "核心受众画像", "二级内容类型", "二级博主身份"]
                           if c in df.columns]
        extra_vals: dict = {}
        if extra_available:
            st.markdown("**扩展筛选**")
            ex_cols = st.columns(len(extra_available))
            for col_ui, field in zip(ex_cols, extra_available):
                opts = ["全部"] + sorted(df[field].dropna().astype(str).unique().tolist())
                extra_vals[field] = col_ui.selectbox(field, opts)

    flt = df.copy()
    if sel_brand != "全部": flt = flt[flt["品牌名称"] == sel_brand]
    if sel_cat   != "全部" and "品类"      in flt.columns: flt = flt[flt["品类"]      == sel_cat]
    if sel_plat  != "全部" and "发布平台"  in flt.columns: flt = flt[flt["发布平台"]  == sel_plat]
    for field, val in extra_vals.items():
        if val != "全部":
            flt = flt[flt[field].astype(str) == val]

    if flt.empty:
        st.warning("当前筛选条件下无数据，请调整")
        st.stop()

    # 核心指标卡（含 help 悬浮公式说明）
    st.markdown("### 📈 核心数据概览")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("内容总数", f"{len(flt)} 篇",
              help="当前筛选条件下的内容总量")
    m2.metric("平均点赞", fmt_num(flt["点赞数"].mean()),
              help="数值已格式化：>= 1w 显示 Xw，>= 1k 显示 Xk")
    m3.metric("平均互动率", f"{flt['互动率'].mean():.4f}",
              help="互动率 = 互动总量 / 博主粉丝量\n"
                   "反映内容在粉丝群体中的渗透效率，排除账号规模干扰")
    m4.metric("最高黑马指数", f"{flt['黑马指数'].max():.2f}",
              help="黑马指数 = 互动率效率 x 0.7 + 点赞规模 x 0.3\n\n"
                   "互动率效率 = 该内容互动率 / 同粉丝层级平均互动率\n"
                   "点赞规模   = 该内容点赞数 / 该品牌平均点赞数\n\n"
                   "指数 > 1.2 视为黑马内容（低粉高流量）")
    m5.metric("黑马内容数", f"{(flt['黑马指数'] > 1.2).sum()} 篇",
              help="黑马指数 > 1.2 的内容数量")
    m6.metric("复合策略占比", f"{(flt['营销动机'] == '复合策略').mean():.1%}",
              help="标题同时命中多个营销动机类别的内容占比\n高复合策略通常代表内容层次更丰富")

    st.markdown("---")

    tab1, tab2, tab3, tab4, tab_insight, tab5 = st.tabs([
        "🏆 爆款内容", "🐴 黑马分析", "⏰ 发布规律", "📊 竞品矩阵", "🎯 人群洞察", "📥 导出报告"
    ])

    # ━━━ Tab1 爆款内容 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with tab1:
        st.subheader("🏆 高互动内容 TOP 10")
        show_cols = ["品牌名称", "内容标题", "营销动机",
                     "点赞数", "收藏数", "评论数", "互动总量", "黑马指数"]
        for extra in ["发布平台", "品类"]:
            if extra in flt.columns:
                show_cols.insert(2, extra)
        top10 = flt.nlargest(10, "互动总量")[show_cols].reset_index(drop=True)
        top10.index += 1
        st.dataframe(top10, use_container_width=True)

        c1, c2 = st.columns(2)
        
        with c1:
            st.subheader("🔑 爆款关键词 TOP 20")
            sample = flt.nlargest(min(50, len(flt)), "互动总量")["内容标题"].tolist()
            kws = extract_keywords(sample)
            if kws:
                kw_viz = pd.DataFrame(kws, columns=["关键词", "出现次数"])
                fig_bar = px.bar(kw_viz, x="出现次数", y="关键词", orientation="h",
                                 title="爆款内容高频词",
                                 color="出现次数", color_continuous_scale="Blues")
                fig_bar.update_layout(yaxis=dict(autorange="reversed"), height=500,
                                      coloraxis_showscale=False)
                st.plotly_chart(fig_bar, use_container_width=True)

        with c2:
            st.subheader("🎯 营销动机分布")
            mot = flt["营销动机"].value_counts().reset_index()
            mot.columns = ["营销动机", "数量"]
            fig_pie = px.pie(mot, names="营销动机", values="数量",
                             title="内容营销动机占比", hole=0.4)
            st.plotly_chart(fig_pie, use_container_width=True)

            st.subheader("📏 动态标题长度 vs 平均互动量")
            bin_step = st.slider("调整字数区间跨度", min_value=1, max_value=5, value=5, step=1, key="title_bin_step", help="左右拖动，寻找不同颗粒度下的爆款标题长度甜区")
            
            # def 函数也是对齐的
            def _dynamic_bin(t_len, step):
                # 只有函数里面的代码，才需要再往右缩进 4 个空格
                if pd.isna(t_len) or t_len == 0: return "[0]"
                low = ((int(t_len) - 1) // step) * step + 1
                return f"[{low}-{low + step - 1}]"
            
            # 下面这些又回到和 st.subheader 一样的对齐线
            flt_len = flt.copy()
            flt_len["当前字数"] = flt_len["内容标题"].astype(str).str.strip().str.len()
            flt_len["动态长度区间"] = flt_len["当前字数"].apply(lambda x: _dynamic_bin(x, bin_step))
            
            len_df = flt_len.groupby("动态长度区间")["互动总量"].mean().reset_index()
            len_df["_sort_key"] = len_df["动态长度区间"].str.extract(r'\[(\d+)').astype(float)
            len_df = len_df.sort_values("_sort_key").drop(columns=["_sort_key"])

            st.plotly_chart(
                px.bar(len_df, x="动态长度区间", y="互动总量",
                       title=f"每 {bin_step} 字区间的平均互动量",
                       color="互动总量", color_continuous_scale="Purples"),
                use_container_width=True
            )

        # ━━━ 新增 Row 2：互动类型拆解 + 品牌 TOP 5 ━━━━━━━━━━━━━━━━━━
        c3, c4 = st.columns(2)
        with c3:
            st.subheader("💬 互动类型拆解")
            interact_cols = ["点赞数", "收藏数", "评论数", "分享数"]
            avail_interact = [c for c in interact_cols if c in flt.columns]
            if avail_interact:
                interact_sums = flt[avail_interact].sum().reset_index()
                interact_sums.columns = ["互动类型", "总量"]
                fig_interact = px.bar(interact_sums, x="互动类型", y="总量",
                                      title="互动类型总量拆解",
                                      color="互动类型",
                                      color_discrete_sequence=px.colors.qualitative.Set2)
                fig_interact.update_layout(showlegend=False, height=420)
                st.plotly_chart(fig_interact, use_container_width=True)

        with c4:
            st.subheader("🏆 品牌 TOP 5 排行")
            brand_top5 = flt.groupby("品牌名称")["互动总量"].sum().nlargest(5).reset_index()
            brand_top5.columns = ["品牌名称", "互动总量"]
            fig_brand5 = px.bar(brand_top5, x="互动总量", y="品牌名称",
                                orientation="h", title="互动总量品牌 TOP 5",
                                color="互动总量", color_continuous_scale="Teal")
            fig_brand5.update_layout(yaxis=dict(autorange="reversed"), height=420,
                                      coloraxis_showscale=False)
            st.plotly_chart(fig_brand5, use_container_width=True)

        # ━━━ 新增 Row 3：发布平台效果对比 ━━━━━━━━━━━━━━━━━━━━━━━━━━━
        if "发布平台" in flt.columns:
            st.subheader("📊 发布平台效果对比")
            plat_metrics = flt.groupby("发布平台").agg(
                平均互动量=("互动总量", "mean"),
                平均互动率=("互动率", "mean"),
                内容数量=("内容标题", "count"),
            ).reset_index()
            c5, c6 = st.columns(2)
            with c5:
                fig_plat_bar = px.bar(plat_metrics, x="发布平台", y="平均互动量",
                                      title="各平台平均互动量",
                                      color="发布平台",
                                      color_discrete_sequence=px.colors.qualitative.Pastel)
                fig_plat_bar.update_layout(showlegend=False, height=380)
                st.plotly_chart(fig_plat_bar, use_container_width=True)
            with c6:
                fig_plat_cnt = px.bar(plat_metrics, x="发布平台", y="内容数量",
                                      title="各平台内容发布量",
                                      color="发布平台",
                                      color_discrete_sequence=px.colors.qualitative.Pastel)
                fig_plat_cnt.update_layout(showlegend=False, height=380)
                st.plotly_chart(fig_plat_cnt, use_container_width=True)

            # ━━━ 新增：AI 爆款一键仿写模块 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        st.markdown("---")
        st.subheader("✨ AI 一键爆款仿写 (从洞察到产出)")
        st.caption("选中一篇爆款，输入自身卖点，让 AI 提取爆款基因并为你生成 3 篇可直接发布的文案。")

        # 1. 交互区：选择爆款 + 输入卖点
        c_ai1, c_ai2 = st.columns([1, 1])
        with c_ai1:
            # 提取 TOP 10 的标题供用户选择
            top_titles = top10["内容标题"].dropna().tolist()
            if top_titles:
                target_title = st.selectbox("🎯 1. 选择你要拆解对标的爆款内容：", top_titles)
                # 获取该爆款的核心基因
                target_row = top10[top10["内容标题"] == target_title].iloc[0]
                motive = target_row.get("营销动机", "未知")
                st.info(f"🧬 该爆款基因提取：**{motive}**")
            else:
                st.warning("暂无爆款数据")
                target_title, motive = "", ""

        with c_ai2:
            my_product = st.text_area("📦 2. 输入我们自己的产品卖点 (Product Brief)：", 
                                      placeholder="例如：我们新推的便携冲牙器，核心卖点是水压大但不伤牙龈，体积像口红一样小，适合上班族饭后使用，首发价 99 元...",
                                      height=130)

        # 2. 触发按钮与 AI Prompt 组装
        if st.button("🚀 召唤 AI 提取基因并生成文案", type="primary", use_container_width=True):
            if not my_product:
                st.warning("⚠️ 请先输入我们自己的产品卖点哦！")
            else:
                with st.spinner("🧠 AI 正在深度拆解爆款逻辑，撰写网感文案中..."):
                    # 组装给大模型的 Prompt（提示词工程）
                    system_prompt = f"""
你是一个百万粉级别的爆款操盘手和文案专家。请深度拆解以下竞品爆款的基因，并结合我的产品卖点，仿写 3 篇可以直接发布的小红书/抖音爆款文案。

【对标爆款基因】
- 爆款原标题：{target_title}
- 营销动机类型：{motive}

【我的产品卖点】
{my_product}

【硬性要求 — 标题多样化】
3 篇文案的标题必须使用完全不同的句式结构，严禁只替换品牌名：
- 第 1 篇：疑问/反常识句（如"为什么XXX？"、"你真的了解XXX吗？"）
- 第 2 篇：数字/对比/清单句（如"用了3个月的真实感受"、"平价vs大牌"）
- 第 3 篇：场景/情感共鸣句（如"上班族必备"、"熬夜党救星"）

【内容要求】
每篇包含：吸睛标题 + 痛点引入 + 产品自然植入（不生硬） + 促单转化结尾。
排版有呼吸感，加入适当 Emoji，语言有网感，像真实博主在写，而非广告腔。
"""
                    
                    if not api_key:
                        st.warning("请先在左侧侧边栏填入 API Key")
                    else:
                        try:
                            result_text = call_ai(api_key, model_type, system_prompt,
                                                  temperature=ai_temperature)
                            st.success("🎉 AI 文案生成完成！")
                            st.markdown(result_text)
                        except Exception as e:
                            st.error(f"AI 调用失败：{e}")

            # ━━━ Tab2 黑马分析 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with tab2:
        st.subheader("🐴 黑马内容排行（黑马指数 Top 10）")
        dh = flt.nlargest(10, "黑马指数")[[
            "品牌名称", "内容标题", "粉丝层级",
            "互动率", "层级平均互动率", "点赞数", "黑马指数", "营销动机"
        ]].reset_index(drop=True)
        dh.index += 1
        st.dataframe(dh, use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            # 去极端值：裁掉 5th–95th 百分位以外的数据，图表更舒展
            bhi_q05 = flt["黑马指数"].quantile(0.05)
            bhi_q95 = flt["黑马指数"].quantile(0.95)
            flt_box = flt[flt["黑马指数"].between(bhi_q05, bhi_q95)]
            fig_box = px.box(flt_box, x="粉丝层级", y="黑马指数",
                             color="粉丝层级", points="outliers",
                             title="各粉丝层级黑马指数分布（动态分层）")
            fig_box.add_hline(y=1.2, line_dash="dash", line_color="red",
                              annotation_text="黑马阈值 1.2")
            st.plotly_chart(fig_box, use_container_width=True)
       
    # ━━━ Tab3 发布规律 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with tab3:
        c1, c2 = st.columns(2)
        with c1:
            st.plotly_chart(
                px.bar(flt.groupby("发布小时")["互动总量"].mean().reset_index(),
                       x="发布小时", y="互动总量", title="各发布时段平均互动量",
                       labels={"发布小时": "发布时段（小时）"},
                       color="互动总量", color_continuous_scale="Oranges"),
                use_container_width=True
            )
        with c2:
            week_cn = {"Monday": "周一", "Tuesday": "周二", "Wednesday": "周三",
                       "Thursday": "周四", "Friday": "周五",
                       "Saturday": "周六", "Sunday": "周日"}
            wk = flt.groupby("发布星期")["互动总量"].mean().reset_index()
            wk["发布星期"] = wk["发布星期"].map(week_cn)
            st.plotly_chart(
                px.bar(wk, x="发布星期", y="互动总量", title="各星期平均互动量",
                       color="互动总量", color_continuous_scale="Greens"),
                use_container_width=True
            )
        daily = flt.groupby("发布日期").size().reset_index(name="发布数量")
        st.plotly_chart(
            px.line(daily, x="发布日期", y="发布数量",
                    title="每日发布频率趋势", markers=True),
            use_container_width=True
        )

    # ━━━ Tab4 竞品矩阵 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with tab4:
        st.subheader("🗺️ 品牌策略四象限：发布规模 vs 内容效率")
        matrix = df.groupby("品牌名称").agg(
            内容发布总量=("内容标题", "count"),
            平均互动率=("互动率", "mean")
        ).reset_index()
        x_mid = matrix["内容发布总量"].median()
        y_mid = matrix["平均互动率"].median()
        x_max = matrix["内容发布总量"].max() * 1.35
        y_max = matrix["平均互动率"].max() * 1.35

        fig_q = go.Figure()
        for x0, x1, y0, y1, color, label in [
            (0,     x_mid, y_mid, y_max, "rgba(255,200,200,0.25)", "潜力型\n量少·效率高"),
            (x_mid, x_max, y_mid, y_max, "rgba(200,255,200,0.25)", "标杆型\n量大·效率高"),
            (0,     x_mid, 0,     y_mid, "rgba(220,220,220,0.25)", "低效型\n量少·效率低"),
            (x_mid, x_max, 0,     y_mid, "rgba(200,220,255,0.25)", "铺量型\n量大·效率低"),
        ]:
            fig_q.add_shape(type="rect", x0=x0, x1=x1, y0=y0, y1=y1,
                            fillcolor=color, line=dict(width=0))
            fig_q.add_annotation(x=(x0+x1)/2, y=(y0+y1)/2, text=f"<b>{label}</b>",
                                 showarrow=False, font=dict(size=13, color="#888"), opacity=0.6)
        fig_q.add_vline(x=x_mid, line_dash="dash", line_color="#aaa", opacity=0.6)
        fig_q.add_hline(y=y_mid, line_dash="dash", line_color="#aaa", opacity=0.6)
        fig_q.add_trace(go.Scatter(
            x=matrix["内容发布总量"], y=matrix["平均互动率"],
            mode="markers+text", text=matrix["品牌名称"], textposition="top center",
            marker=dict(size=20, color="royalblue", opacity=0.85,
                        line=dict(width=2, color="white")),
            hovertemplate="<b>%{text}</b><br>发布量:%{x}<br>平均互动率:%{y:.6f}<extra></extra>"
        ))
        fig_q.update_layout(title="品牌策略四象限（X=发布规模 · Y=内容效率）",
                            xaxis_title="内容发布总量", yaxis_title="平均互动率",
                            height=580, showlegend=False)
        st.plotly_chart(fig_q, use_container_width=True)
        st.info("标杆型：量大高效 | 潜力型：质量好可加码 | 铺量型：靠量取胜 | 低效型：需调整策略")

        c1, c2 = st.columns(2)
        with c1:
            bp = flt.groupby("品牌名称")[
                ["点赞数", "收藏数", "评论数", "分享数"]
            ].mean().reset_index()
            st.plotly_chart(
                px.bar(bp.melt(id_vars="品牌名称", var_name="指标", value_name="平均值"),
                       x="品牌名称", y="平均值", color="指标", barmode="group",
                       title="各品牌互动指标对比"),
                use_container_width=True
            )
        with c2:
            if "博主类型" in flt.columns:
                bt = flt.groupby(["品牌名称", "博主类型"]).size().reset_index(name="数量")
                st.plotly_chart(
                    px.bar(bt, x="品牌名称", y="数量", color="博主类型",
                           title="各品牌博主类型策略", barmode="stack"),
                    use_container_width=True
                    )
            # ━━━ 新增 Tab: 人群洞察 (交叉热力图) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        with tab_insight:
            st.subheader("🔥 受众 x 营销动机：交叉转化热力图")
            st.caption("颜色越深，代表该类受众对特定营销动机的转化效率越高。")

            # 1. 动态获取当前数据中可用的受众维度
            available_dims = []
            if "粉丝年龄段" in flt.columns: available_dims.append("粉丝年龄段")
            if "核心受众画像" in flt.columns: available_dims.append("核心受众画像")
            if "博主类型" in flt.columns: available_dims.append("博主类型") # 作为备用维度

            if not available_dims:
                st.info("👀 当前数据源缺少受众相关字段（如：粉丝年龄段、核心受众画像），无法生成热力图。")
            else:
                # 2. 交互式选择器：让运营自己决定看什么
                col_a, col_b = st.columns(2)
                cross_dim = col_a.selectbox("选择人群维度 (Y轴)", available_dims)
                val_metric = col_b.selectbox("选择衡量指标 (决定颜色深浅)", ["平均黑马指数", "互动率", "平均互动量"])

                # 映射用户选择到实际列名
                metric_map = {"互动率": "互动率", "平均黑马指数": "黑马指数", "平均互动量": "互动总量"}
                target_col = metric_map[val_metric]

                # 3. Pandas 数据透视 (核心逻辑)
                heatmap_data = pd.pivot_table(
                    flt,
                    values=target_col,
                    index=cross_dim,
                    columns="营销动机",
                    aggfunc="mean"
                ).fillna(0) # 填补空缺数据

                # 4. 绘制 Plotly 热力图
                if not heatmap_data.empty:
                    # 使用不同色系区分不同指标的业务感
                    color_scale = "Reds" if val_metric == "平均互动量" else "Teal" 
                    
                    fig_hm = px.imshow(
                        heatmap_data,
                        text_auto=".2f" if val_metric != "平均互动量" else ".0f", # 自动在格子里显示数值
                        aspect="auto",
                        color_continuous_scale=color_scale,
                        labels=dict(color=val_metric)
                    )
                    fig_hm.update_layout(
                        xaxis_title="营销动机 (套路)",
                        yaxis_title=cross_dim,
                        height=500
                    )
                    st.plotly_chart(fig_hm, use_container_width=True)

                    # 5. PM 的小巧思：自动输出一条行动建议
                    try:
                        best_combination = heatmap_data.stack().idxmax()
                        max_val = heatmap_data.stack().max()
                        fmt_val = f"{max_val:.2f}" if val_metric != "平均互动量" else fmt_num(max_val)
                        st.success(
                            f"💡 **AI 策略提示**：当前数据下，**「{best_combination[0]}」**群体最吃**「{best_combination[1]}」**这一套！"
                            f"其 {val_metric} 飙升至 **{fmt_val}**，建议将预算向该组合倾斜。"
                        )
                    except Exception:
                        pass

    # ━━━ Tab5 导出报告 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with tab5:
        strategy_facts = build_strategy_facts(flt)

        st.subheader("📋 Strategy_Facts 预览")
        st.dataframe(strategy_facts, use_container_width=True)

        st.markdown("---")
        st.subheader("🤖 AI 洞察报告")

        if not api_key:
            st.warning("请先在侧边栏输入 API Key")
        else:
            if st.button("✨ 一键生成 AI 洞察报告",
                         use_container_width=True, type="primary"):
                with st.spinner("🤖 AI 分析中（10-30 秒）…"):
                    try:
                        st.session_state.ai_insight = call_ai(
                            api_key, model_type, format_prompt(strategy_facts, flt),
                            temperature=ai_temperature,
                        )
                    except Exception as e:
                        msg = str(e)
                        if "timeout"   in msg.lower(): st.error("请求超时，请检查网络")
                        elif "429"     in msg:         st.error("API 频率限制，请等待 1 分钟")
                        elif "401"     in msg:         st.error("API Key 无效")
                        else:                          st.error(f"生成失败：{msg}")

        if st.session_state.ai_insight:
            st.markdown("### 📝 AI 分析结论")
            st.markdown(st.session_state.ai_insight)

        st.markdown("---")
        st.subheader("📥 导出报告（含 Excel 图表 / 首行冻结 / 色阶 / 自动筛选）")

        sample_titles = flt.nlargest(min(50, len(flt)), "互动总量")["内容标题"].tolist()
        kw_exp    = pd.DataFrame(extract_keywords(sample_titles), columns=["关键词", "出现次数"])
        top10_exp = flt.nlargest(10, "互动总量")[[
            "品牌名称", "内容标题", "营销动机",
            "点赞数", "收藏数", "评论数", "互动总量", "黑马指数"
        ]]
        brand_exp = flt.groupby("品牌名称").agg(
                    内容发布量=("内容标题", "count"),  # 新增：供散点图 X 轴使用
                    点赞数=("点赞数", "mean"),
                    收藏数=("收藏数", "mean"),
                    评论数=("评论数", "mean"),
                    互动总量=("互动总量", "mean"),
                    互动率=("互动率", "mean"),        # 供散点图 Y 轴使用
                    黑马指数=("黑马指数", "mean")
                ).round(4).reset_index()

        # 新增导出数据：互动类型 / 品牌 TOP5 / 平台对比
        interact_cols = ["点赞数", "收藏数", "评论数", "分享数"]
        avail_interact = [c for c in interact_cols if c in flt.columns]
        interact_exp = flt[avail_interact].sum().reset_index()
        interact_exp.columns = ["互动类型", "总量"]

        brand_top5_exp = flt.groupby("品牌名称")["互动总量"].sum().nlargest(5).reset_index()
        brand_top5_exp.columns = ["品牌名称", "互动总量"]

        platform_exp = pd.DataFrame()
        if "发布平台" in flt.columns:
            platform_exp = flt.groupby("发布平台").agg(
                平均互动量=("互动总量", "mean"),
                平均互动率=("互动率", "mean"),
                内容数量=("内容标题", "count"),
            ).round(4).reset_index()

        excel_out = build_excel(
            flt, strategy_facts, top10_exp, kw_exp,
            brand_exp, st.session_state.ai_insight,
            interact_df=interact_exp,
            brand_top5_df=brand_top5_exp,
            platform_df=platform_exp,
        )
        today = datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            "📥 下载完整分析报告（Excel · 9 Sheet）",
            data=excel_out.getvalue(),
            file_name=f"竞品分析报告_{today}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.markdown(
            "**包含**：原始数据 · TOP10爆款 · 关键词分析 · "
            "竞品对比(嵌入图表) · Strategy_Facts(色阶) · 互动类型拆解 · 品牌TOP5 · 平台效果对比 · AI_Insight\n\n"
            "**格式**：首行冻结 · 自动筛选 · 深蓝表头 · 自动列宽"
        )
